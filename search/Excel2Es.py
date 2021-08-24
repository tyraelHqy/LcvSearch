from openpyxl import load_workbook
from search.models import QAType

from elasticsearch_dsl.connections import connections

es = connections.create_connection(QAType._doc_type.using)


ANSWER_IS_NONE = '不好意思，无法回答。'
URL_IS_NONE = 'https://doc.weixin.qq.com/txdoc/excel?scode=AJEAIQdfAAo4sb50S1AJcA2wZ-ACc&docid=e2_AJcA2wZ-ACcmB6iHRxjRNq4uR8PlX&type=1'
number = 88
filename = "外部对接问题收敛.xlsx"
tablename = "工作表1"
column_name = "问题描述"


def gen_suggests(index, info_tuple):
    # 根据字符串生成搜索建议数组
    used_words = set()
    suggests = []
    for text, weight in info_tuple:
        if text:
            # 调用es的analyze接口分析字符串
            words = es.indices.analyze(index=index, analyzer="ik_max_word", params={'filter': ["lowercase"]}, body=text)
            analyzed_words = set([r["token"] for r in words["tokens"] if len(r["token"]) > 1])
            new_words = analyzed_words - used_words
        else:
            new_words = set()
        if new_words:
            suggests.append({"input": list(new_words), "weight": weight})

    return suggests


class QaItem(object):
    def __init__(self, question, answer, env, url, indexId):
        self.question = question
        self.answer = answer
        self.env = env
        self.url = url
        self.indexId = indexId


def getExcelData(name, number_of_excel):
    # 1.打开 Excel 表格并获取表格名称
    questionList = []
    answerList = []
    envList = []
    urlList = []
    qaList = []
    workbook = load_workbook(filename="../static/DBFile/" + name)
    ws = workbook.get_sheet_by_name(tablename)
    question_column = ws['E']
    answer_column = ws['F']
    env_column = ws['D']
    url_column = ws['G']
    for x in range(number_of_excel):
        questionList.append(question_column[x].value)
        if answer_column[x].value is not None:
            answer = answer_column[x].value
        else:
            answer = ANSWER_IS_NONE

        if url_column[x].value is not None:
            url = url_column[x].value
        else:
            url = URL_IS_NONE

        answerList.append(answer)
        envList.append(env_column[x].value)
        urlList.append(url_column[x].value)
        qaItem = QaItem(question_column[x].value, answer, env_column[x].value, url, x)
        qaList.append(qaItem)
    qaList.pop(0)
    return qaList


def saveType(qa_dict):
    for item in qa_dict:
        # 将dict转成es的数据
        qaType = QAType()
        qaType.question = item.question
        qaType.answer = item.answer
        qaType.url = item.url
        qaType.env = item.env
        qaType.meta.id = item.indexId

        qaType.suggest = gen_suggests(QAType._doc_type.index, ((qaType.question, 10), (qaType.answer, 3),))

        qaType.save()


def run():
    qa_dict = getExcelData(filename, number)
    print(len(qa_dict))
    saveType(qa_dict)


if __name__ == "__main__":
    run()
